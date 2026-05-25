from __future__ import annotations

import base64
import html
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import openpyxl
import requests

from ..manifest import append_manifest, manifest_failure, manifest_for_file, source_by_id
from ..normalize import version_from_text
from ..paths import path, relpath


ARCTICGRO_CITATION = "The Arctic Great Rivers Observatory; cite the dataset and Version YYYYMMDD from https://www.arcticgreatrivers.org/data/."
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
XLSX_MIME_MARKER = "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._current_href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "a":
            attr = dict(attrs)
            self._current_href = attr.get("href")
            self._text = []

    def handle_data(self, data: str) -> None:
        if self._current_href:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._current_href:
            text = " ".join(" ".join(self._text).split())
            self.links.append((text, self._current_href))
            self._current_href = None
            self._text = []


@dataclass
class DownloadedText:
    text: str
    url: str
    resolved_url: str
    local_path: Path


def _request(url: str, *, timeout: int = 90) -> requests.Response:
    response = requests.get(url, headers=REQUEST_HEADERS, timeout=timeout, allow_redirects=True)
    response.raise_for_status()
    return response


def _write_bytes(destination: Path, content: bytes) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(content)


def _write_text(destination: Path, text: str) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(text, encoding="utf-8")


def _discover_links(page_url: str, html_text: str) -> dict[str, str]:
    parser = LinkParser()
    parser.feed(html_text)
    out: dict[str, str] = {}
    for text, href in parser.links:
        if not href:
            continue
        absolute = urljoin(page_url, href)
        out[text.strip().lower()] = absolute
    return out


def _find_link(links: dict[str, str], contains: str, fallback: str) -> str:
    needle = contains.lower()
    for text, href in links.items():
        if needle in text:
            return href
    return fallback


def _google_sheet_id(url: str) -> str | None:
    match = re.search(r"/spreadsheets/d/([^/]+)/", url)
    return match.group(1) if match else None


def _google_export_url(url: str, fmt: str = "xlsx") -> str:
    sheet_id = _google_sheet_id(url)
    if not sheet_id:
        raise ValueError(f"Not a Google Sheets URL: {url}")
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format={fmt}"


def _drive_file_id(url: str) -> str | None:
    match = re.search(r"/file/d/([^/]+)/", url)
    return match.group(1) if match else None


def _drive_download_url(url: str) -> str:
    file_id = _drive_file_id(url)
    if not file_id:
        return url
    return f"https://drive.google.com/uc?export=download&id={file_id}"


def _detect_xlsx_version(workbook_path: Path) -> str:
    try:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        texts: list[str] = []
        for sheet_name in workbook.sheetnames[:3]:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
                texts.extend(str(value) for value in row if value is not None)
        workbook.close()
        return version_from_text(*texts)
    except Exception:
        return version_from_text(workbook_path.name)


def _download_binary(
    *,
    source_id: str,
    source_url: str,
    download_url: str,
    destination: Path,
    version_hint: str = "",
    dry_run: bool = False,
) -> None:
    if dry_run:
        append_manifest(
            {
                "source_id": source_id,
                "source_url": source_url,
                "resolved_url": "",
                "download_url": download_url,
                "retrieved_at_utc": "",
                "local_path": relpath(destination),
                "file_name": destination.name,
                "file_size_bytes": "",
                "sha256": "",
                "version_detected": version_hint,
                "download_status": "dry_run",
                "failure_reason": "",
                "license_or_citation": ARCTICGRO_CITATION,
                "commit_raw_data": False,
            }
        )
        return
    try:
        response = _request(download_url)
        _write_bytes(destination, response.content)
        version = _detect_xlsx_version(destination) or version_hint or version_from_text(destination.name)
        append_manifest(
            manifest_for_file(
                source_id=source_id,
                source_url=source_url,
                download_url=download_url,
                resolved_url=response.url,
                local_path=destination,
                version_detected=version,
                license_or_citation=ARCTICGRO_CITATION,
            )
        )
    except Exception as exc:
        append_manifest(
            manifest_failure(
                source_id=source_id,
                source_url=source_url,
                download_url=download_url,
                failure_reason=str(exc),
                local_path=destination,
                license_or_citation=ARCTICGRO_CITATION,
                version_detected=version_hint,
            )
        )


def _download_text_page(
    *,
    source_id: str,
    source_url: str,
    destination: Path,
    dry_run: bool = False,
) -> DownloadedText | None:
    if dry_run:
        append_manifest(
            {
                "source_id": source_id,
                "source_url": source_url,
                "resolved_url": "",
                "download_url": source_url,
                "retrieved_at_utc": "",
                "local_path": relpath(destination),
                "file_name": destination.name,
                "file_size_bytes": "",
                "sha256": "",
                "version_detected": "",
                "download_status": "dry_run",
                "failure_reason": "",
                "license_or_citation": ARCTICGRO_CITATION,
                "commit_raw_data": False,
            }
        )
        return None
    try:
        response = _request(source_url)
        _write_text(destination, response.text)
        version = version_from_text(response.text[:10000], destination.name)
        append_manifest(
            manifest_for_file(
                source_id=source_id,
                source_url=source_url,
                download_url=source_url,
                resolved_url=response.url,
                local_path=destination,
                version_detected=version,
                license_or_citation=ARCTICGRO_CITATION,
            )
        )
        return DownloadedText(response.text, source_url, response.url, destination)
    except Exception as exc:
        append_manifest(
            manifest_failure(
                source_id=source_id,
                source_url=source_url,
                download_url=source_url,
                failure_reason=str(exc),
                local_path=destination,
                license_or_citation=ARCTICGRO_CITATION,
            )
        )
        return None


def _manifest_manual_required(source_id: str, reason: str) -> None:
    record = source_by_id(source_id)
    append_manifest(
        {
            "source_id": source_id,
            "source_url": record["source_url"],
            "resolved_url": record["source_url"],
            "download_url": record["source_url"],
            "retrieved_at_utc": "",
            "local_path": "",
            "file_name": "",
            "file_size_bytes": "",
            "sha256": "",
            "version_detected": "",
            "download_status": "manual_required",
            "failure_reason": reason,
            "license_or_citation": ARCTICGRO_CITATION,
            "commit_raw_data": False,
        }
    )


def _extract_embedded_discharge_workbooks(html_text: str, dry_run: bool = False) -> None:
    record = source_by_id("arcticgro_discharge_current")
    workbook_map: dict[str, str] = record.get("discharge_workbooks", {})
    raw_dir = path(record.get("local_subdir", "data/raw/arcticgro/discharge"))
    source_url = record["source_url"]
    for river, file_name in workbook_map.items():
        out_path = raw_dir / file_name
        if dry_run:
            append_manifest(
                {
                    "source_id": "arcticgro_discharge_current",
                    "source_url": source_url,
                    "resolved_url": source_url,
                    "download_url": f"embedded:{file_name}",
                    "retrieved_at_utc": "",
                    "local_path": relpath(out_path),
                    "file_name": file_name,
                    "file_size_bytes": "",
                    "sha256": "",
                    "version_detected": version_from_text(file_name),
                    "download_status": "dry_run",
                    "failure_reason": "",
                    "license_or_citation": ARCTICGRO_CITATION,
                    "commit_raw_data": False,
                }
            )
            continue
        try:
            idx = html_text.find(file_name)
            if idx < 0:
                raise ValueError(f"Workbook filename not found in discharge page: {file_name}")
            marker_idx = html_text.rfind(XLSX_MIME_MARKER, 0, idx)
            if marker_idx < 0:
                raise ValueError(f"Embedded xlsx payload marker not found for {file_name}")
            start = marker_idx + len(XLSX_MIME_MARKER)
            end_candidates = [
                candidate
                for candidate in [
                    html_text.find("&#39;);", start),
                    html_text.find("');", start),
                    html_text.find('"', start),
                ]
                if candidate > start
            ]
            if not end_candidates:
                raise ValueError(f"Could not locate end of embedded payload for {file_name}")
            payload = base64.b64decode(html.unescape(html_text[start : min(end_candidates)]).strip())
            _write_bytes(out_path, payload)
            append_manifest(
                manifest_for_file(
                    source_id="arcticgro_discharge_current",
                    source_url=source_url,
                    download_url=f"embedded:{file_name}",
                    resolved_url=source_url,
                    local_path=out_path,
                    version_detected=version_from_text(file_name),
                    license_or_citation=ARCTICGRO_CITATION,
                )
            )
        except Exception as exc:
            append_manifest(
                manifest_failure(
                    source_id="arcticgro_discharge_current",
                    source_url=source_url,
                    download_url=f"embedded:{file_name}",
                    failure_reason=f"{river}: {exc}",
                    version_detected=version_from_text(file_name),
                    local_path=out_path,
                    license_or_citation=ARCTICGRO_CITATION,
                )
            )


def download_arcticgro(dry_run: bool = False) -> None:
    data_record = source_by_id("arcticgro_data_page")
    data_page = _download_text_page(
        source_id="arcticgro_data_page",
        source_url=data_record["source_url"],
        destination=path(data_record.get("local_subdir", "data/raw/arcticgro")) / "arcticgro_data_page.html",
        dry_run=dry_run,
    )
    links: dict[str, str] = {}
    if data_page:
        links = _discover_links(data_page.resolved_url, data_page.text)

    water_record = source_by_id("arcticgro_water_quality_current")
    water_url = _find_link(links, "Water Quality", water_record["source_url"])
    water_page = _download_text_page(
        source_id="arcticgro_water_quality_current",
        source_url=water_url,
        destination=path(water_record.get("local_subdir", "data/raw/arcticgro/water_quality")) / "arcticgro_water_quality_page.html",
        dry_run=dry_run,
    )
    water_links = _discover_links(water_url, water_page.text) if water_page else {}
    water_sheet_url = _find_link(water_links, "Google Sheets", water_record["source_url"])
    try:
        water_export_url = _google_export_url(water_sheet_url, "xlsx")
    except ValueError:
        water_export_url = _google_export_url(water_record["source_url"], "xlsx")
    _download_binary(
        source_id="arcticgro_water_quality_current",
        source_url=water_record["source_url"],
        download_url=water_export_url,
        destination=path(water_record.get("local_subdir", "data/raw/arcticgro/water_quality")) / "arcticgro_water_quality_current.xlsx",
        dry_run=dry_run,
    )

    param_record = source_by_id("arcticgro_water_quality_parameter_descriptions")
    _download_binary(
        source_id="arcticgro_water_quality_parameter_descriptions",
        source_url=param_record["source_url"],
        download_url=_google_export_url(param_record["source_url"], "xlsx"),
        destination=path(param_record.get("local_subdir", "data/raw/arcticgro/metadata")) / "arcticgro_parameter_descriptions.xlsx",
        dry_run=dry_run,
    )

    metadata_record = source_by_id("arcticgro_water_quality_metadata")
    _download_binary(
        source_id="arcticgro_water_quality_metadata",
        source_url=metadata_record["source_url"],
        download_url=_drive_download_url(metadata_record["source_url"]),
        destination=path(metadata_record.get("local_subdir", "data/raw/arcticgro/metadata")) / "arcticgro_metadata.pdf",
        dry_run=dry_run,
    )

    absorbance_record = source_by_id("arcticgro_absorbance_current")
    absorbance_url = _find_link(links, "Absorbance", absorbance_record["source_url"])
    _download_binary(
        source_id="arcticgro_absorbance_current",
        source_url=absorbance_record["source_url"],
        download_url=_google_export_url(absorbance_url, "xlsx"),
        destination=path(absorbance_record.get("local_subdir", "data/raw/arcticgro/absorbance")) / "arcticgro_absorbance_current.xlsx",
        dry_run=dry_run,
    )

    discharge_record = source_by_id("arcticgro_discharge_current")
    discharge_url = _find_link(links, "Discharge", discharge_record["source_url"])
    discharge_page = _download_text_page(
        source_id="arcticgro_discharge_current",
        source_url=discharge_url,
        destination=path(discharge_record.get("local_subdir", "data/raw/arcticgro/discharge")) / "arcticgro_discharge_page.html",
        dry_run=dry_run,
    )
    _extract_embedded_discharge_workbooks(discharge_page.text if discharge_page else "", dry_run=dry_run)

    for source_id, reason in [
        ("arcticgro_water_quality_archived", "Google Drive archive folder is not bulk-downloaded automatically."),
        ("arcticgro_absorbance_archived", "Google Drive archive folder is not bulk-downloaded automatically."),
        ("arcticgro_discharge_archived", "Google Drive archive folder is not bulk-downloaded automatically."),
        ("arcticgro_spatial_data", "Spatial folder may contain multiple files; manual review required before download."),
        ("arcticgro_water_quality_flags", "Flag codes are extracted during preprocessing from the Water Quality workbook."),
    ]:
        _manifest_manual_required(source_id, reason)

